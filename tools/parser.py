from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelParser:

    def __init__(self, filename: Path):

        self.filename = filename

        self.workbook = None

    def load(self):

        self.workbook = load_workbook(
            self.filename,
            data_only=True
        )

    def worksheet_names(self) -> list[str]:

        return self.workbook.sheetnames

    def worksheet(self, name: str) -> Worksheet:

        return self.workbook[name]